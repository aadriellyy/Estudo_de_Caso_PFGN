from datetime import datetime
from pytz import timezone
from openpyxl import load_workbook
import pandas as pd 


#Criando um DataFrame vazio:

# Carregar o arquivo Excel
caminho_arquivo = r"C:\Users\limaa\OneDrive\Área de Trabalho\processo_cogep\Estudo_de_Caso_PGFN.xlsx" #carregando o arquivo, mas isso nao faz com que ele possa ser lido
arquivo_excel = load_workbook(caminho_arquivo) #load_workbook é uma funcao do python, mais propriamente dito da biblioteca openpyxl, usada para possibilitar a leitura da planilha

# Acessar cada planilha
servidores = arquivo_excel['Servidores'] 

#Definindo o maximo de linhas e colunas da planilha seervidores
maxRow = servidores.max_row
maxColumn = servidores.max_column

#Definindo as variaveis que serao usadas posteriormente
ano_atual = 2024
aposentavel = False

#convertendo o codigo da data
result = [] #nessa lista irei armazenar uma sublista que guarda todos os dados de cada usuario
for i in range(2, maxRow+1):
    
    subList = [] #essa eh a sublista usada para armazenar os dados requeridos de cada usuario

    cell_date = servidores.cell(row = i, column = 8) #definindo a data
    cell_sexo = servidores.cell(row = i, column = 6) #definindo o sexo 
    cell_name = servidores.cell(row = i, column = 10) #definindo o nome
        
    timestamp = float(cell_date.value) #atribuindo a variavel timestamp o valor da data extraida da planilha e transformando esse valor em um numero decimal

    # converter o timestamp para uma data e hora em um timezone específico
    dt = datetime.fromtimestamp(timestamp, tz = timezone("America/Sao_Paulo")) #a variavel dt armazena a conversao do timestamp para o fuso-horario brasileiro
    idade = ano_atual - int(dt.year) #definindo a idade do usuario de acordo com o ano atual
    
    #condicoes para receber a aposentadoria
    if idade > 60 and cell_sexo.value == 'F': #para o caso em que o usuario seja do sexo feminino e aposentavel
        aposentavel = True
    elif idade > 65 and cell_sexo.value == 'M': #para o caso em que o usuario seja do sexo masculino e aposentavel
        aposentavel = True
    else:
        aposentavel = False #para o caso em que o usuario nao eh aposentavel
        
    #adicionando a sublista todos os dados necessarios para cada usuario
    subList.append(cell_name.value)
    subList.append(cell_sexo.value)
    subList.append(idade)
    subList.append(aposentavel)
    
    #adicionando a sublista na lista geral citada anteriormente
    result.append(subList)
        

#definindo um DataFrame em que os dados estao contidos na lista result e cada um deve ser separado de acordo com os parametros que estao em columns
df = pd.DataFrame(result, columns = ['Nome', 'Sexo', 'Idade', 'Aposentavel'])
print(df)