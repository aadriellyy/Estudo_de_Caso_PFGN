from openpyxl import load_workbook

# Carregar o arquivo Excel
caminho_arquivo = r"C:\Users\limaa\OneDrive\Área de Trabalho\processo_cogep\Estudo_de_Caso_PGFN.xlsx" #carregando o arquivo, mas isso nao faz com que ele possa ser lido
arquivo_excel = load_workbook(caminho_arquivo) #load_workbook é uma funcao do python, mais propriamente dito da biblioteca openpyxl, usada para possibilitar a leitura da planilha

# Acessar cada planilha
servidores = arquivo_excel['Servidores'] 
unidade = arquivo_excel['Unidade']

#Definindo o maximo de linhas e colunas da planilha seervidores
maxRow = servidores.max_row
maxColumn = servidores.max_column

#Definindo o maximo de linhas da planilha unidade
maxRowUni = unidade.max_row

#inicializando um dicionario com a sigla de todas as UF'S (feito de forma manual para nao sobrecarregar o tempo de execucao)
uf = dict.fromkeys(['AC', 'AL', 'AP', 'AM', 'BA', 'CE', 'DF', 'ES', 'GO', 'MA', 'MT', 'MS', 'MG', 'PA', 'PB', 'PR', 'PE', 'PI', 'RJ', 'RN', 'RS', 'RO', 'RR', 'SC', 'SP', 'SE', 'TO'],0)

#definindo quantos servidores existem por uf
for i in range(1, maxRow + 1):
    cell_obj = servidores.cell(row = i, column = 36) #a varriavel cell_obj armazena a celula na linha i (muda conforme o loop avança) na coluna 36 (coluna AJ)
    for j in range(1, maxRowUni): 
        cell_objUni = unidade.cell(row = j, column = 3) #aqui armazeno o codigo correspondente a cada uf na planilha unidade
        if cell_obj.value == cell_objUni.value: #para cada um dos codigos da coluna AJ eu busco na planilha unidades o codigo que seja igual, assim definindo a uf correspondente
            cell_name = unidade.cell(row = j, column = 9)
            if cell_name.value in uf: #esse tratamento é para os casos em que a uf nao esta listada. se ela nao esta no dicionario, entao nao se trata de uma uf valida. eh apenas um simples tratamento para possiveis erros
                uf[cell_name.value] += 1
print(uf)

#verificando o total de mulheres:
womens = 0 #define a qtd de mulheres
man = 0 #define a qtd de homens
for i in range(1, maxRow +1):
    cell_obj = servidores.cell(row = i, column = 6) #a variavel cell_obj armazena o valor indicado por cada usuario na coluna 6 (codsexo)
    if cell_obj.value == 'F': #caso o usuario se identifique como genero feminino, vou iterando a variavel que define a quantidade de mulheres
        womens += 1
    else:
        man += 1 #caso contrario, como so a duas opcoes nessa planilha, itero a variavel que define a quantidade de homens
total = man+womens #definindo o total de servidores. isso eh util para definir a porcentagem de mulheres que sao servidoras
print(f'Total de servidores: {total}')
print(f'Mulheres servidoras: {womens}')

#verificando o total de servidores de origem AGU com situação funcional EXERC DESCENT CARREI
total_serv = 0 #variavel que define o total de servidores de origem AGU com os criterios da situacao funcional
for i in range (1, maxRow+1):
    cell_orig = servidores.cell(row = i, column = 73) #a variavel cell_orig esta armazenando o orgao de origem do servidor (coluna BU)
    cell_func = servidores.cell(row = i, column = 68) #a variavel cell_func esta armazenando a funcao exercida por esse servidor (coluna BP)
    if cell_orig.value == 'AGU' and 'EXERC. DESCENT.' in cell_func.value: #verificando se as condicoes sao cumrpidas
        total_serv += 1 #caso sejam, itero o total de servidores

print(f'Servidores de origem AGU com situação funcional EXERC DESCENT CARREI: {total_serv}')
